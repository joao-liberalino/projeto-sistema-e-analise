#importa bibliotecas 
from openpyxl   import load_workbook
from tkinter import *
import os
from openpyxl import Workbook
from tkinter import messagebox


# Verifica se o arquivo existe
arquivo_excel = "PROJETO DISBRA1.xlsx"

if not os.path.exists(arquivo_excel):
    print("Arquivo não encontrado. Criando novo arquivo...")
    wb = Workbook()
    wb.save(arquivo_excel)
else:
    wb = load_workbook(arquivo_excel)

#agrupar dados 
dados = ()
#criar codigo de cadastro 
wb = load_workbook("PROJETO DISBRA1.xlsx")
ws = wb.active
#achar linha vazia para adicionar cadstro
linha = ws.max_row + 1 
#escrever os dados 
for i, valor in enumerate(dados, start=1):
    ws.cell(row=linha, column=i, value=valor)
#salvar dados
wb.save("PROJETO DISBRA1.xlsx")
print("cadastro salvo com sucesso  ")

#front end
#utilizar tkinter 

#CRIAR CAMINHO PARA A PLANILHA 

caminho_planilha = "PROJETO DISBRA1.xlsx"

#VERIFICAR SE A PLANILHA EXISTE 
if not os.path.exists(caminho_planilha):
    wb = Workbook()
    ws = wb.active
    ws.append(["cnpj","nome","fone","cep","endereço","bairro", "municipio","uf", "data_emissao", "pat do tanque", "litragem", "pat da bomba","modelo", "pat da bacia"])
    wb.save(caminho_planilha)

#começar a estrutura front 
app = Tk()
app.attributes("-fullscreen", True)
app.bind("<Escape>", lambda e: app.attributes("-fullscreen", False))
app.title("cadastro de cliente disbra diesel")#titulo da interface

labels = ["CNPJ", "Nome", "Fone", "CEP", "Endereço", "Bairro", "Município", "UF",
          "Data de Emissão", "PAT do Tanque", "Litragem", "PAT da Bomba", "Modelo", "PAT da Bacia"]
#criar uma lista com nome dos campos e um dicionario (entries) para armazenar os campos de entradas 

entries = {}
#criar label com nome do campo
#criar um entry pro usuario digitar os dados 
#salvar o entry no dicionario entries, com uma chave simples (ex;"data_emissao")
entries = {}
campos_ordenados = []  # Lista para manter a ordem dos Entry

for i, label_text in enumerate(labels):
    Label(app, text=label_text + ":").grid(row=i, column=0, sticky="e", padx=5, pady=2)
    entry = Entry(app, width=40)
    entry.grid(row=i, column=1, padx=5, pady=2)
    entries[label_text.lower().replace(" ", "_")] = entry
    campos_ordenados.append(entry)

# Funções para avançar/voltar com Tab e Shift+Tab
def foco_proximo(event, i):
    if i + 1 < len(campos_ordenados):
        campos_ordenados[i + 1].focus_set()
    return "break"  # Impede o comportamento padrão

def foco_anterior(event, i):
    if i - 1 >= 0:
        campos_ordenados[i - 1].focus_set()
    return "break"

# Vincular eventos de Tab e Shift+Tab
for i, entry in enumerate(campos_ordenados):
    entry.bind("<Tab>", lambda e, i=i: foco_proximo(e, i))
    entry.bind("<Shift-Tab>", lambda e, i=i: foco_anterior(e, i))



#criar uma funçao 
def salvar():
    dados = [entry.get() for entry in entries.values()] #pegar todos os dados digitados e guarda numa lista chamada dados 
#verificar se todos os campos estao prenchidos 
    if all(dados): # so proseeguir se todos os campos estiver prenchidos 
      wb = load_workbook(caminho_planilha)  
      ws = wb.active
#procurar a proxima linha vazia onde o valor sera inserido      
    linha = 2
    while any(ws.cell(row=linha, column=col).value for col in range(1, len(dados) + 1)): # garantir que os dados vai ser escrito na proxima linha vazia 
         linha += 1
#escreve cada valor da lista dados nas colunas das listas encontradas         
    for i, valor in enumerate(dados, start=1):
         ws.cell(row=linha, column=i, value=valor)
#salvar as alteraçoes no execel e mostrar uma mensagem de sucesso 
         wb.save(caminho_planilha)

         messagebox.showinfo("sucesso", f"cadastro salvo na linha{linha}") 

#limpar campos
    for entry in entries.values():
     entry.delete(0, 'end')

    else:
      
      messagebox.showwarning("Erro", "Preencha todos os campos.")

# criando botton para salvar 
Button(app, text="Salvar Cadastro", command=salvar).grid(row=len(labels), column=0, columnspan=2, pady=1)
app.bind('<Return>', lambda event: salvar())

app.mainloop()
